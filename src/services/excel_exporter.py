import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from models.comparison_result import ComparisonResult


class ExcelExporter:
    def __init__(self, results: list[ComparisonResult]):
        self.results = results

    def export(self, output_path: str):
        data = []

        for r in self.results:
            data.append({
                "Search Value": r.search_value,
                "Status": r.status,
                "File 1 Result": r.file1_value,
                "File 2 Result": r.file2_value,
                "Mismatched Fields": ", ".join(r.mismatch_fields),
            })

        df = pd.DataFrame(data)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Comparison Report")

        wb = load_workbook(output_path)
        ws = wb.active

        GREEN = PatternFill(start_color="C6EFCE", fill_type="solid")
        RED = PatternFill(start_color="FFC7CE", fill_type="solid")
        YELLOW = PatternFill(start_color="FFEB9C", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            status_cell = ws[f"B{row}"]

            if status_cell.value == "MATCH":
                status_cell.fill = GREEN
            elif status_cell.value == "MISMATCH":
                status_cell.fill = RED
            else:
                status_cell.fill = YELLOW

        wb.save(output_path)
